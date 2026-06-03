import os
from flask import Flask, render_template, request, redirect, session, send_file
from functools import wraps
import pandas as pd
import datetime
import psycopg2
import psycopg2.extras
from psycopg2 import pool
from dotenv import load_dotenv
load_dotenv()
import io

app = Flask(__name__)

# ============================================================
# CREDENTIALS FROM .env  (never hardcode in production)
# In your .env file add:
#   SECRET_KEY=some_random_secret_here
#   LOGIN_USERNAME=6301403729
#   LOGIN_PASSWORD=Dad99128@
# ============================================================
app.secret_key = os.getenv("SECRET_KEY", "cement_store_secret_key_99128")
LOGIN_USERNAME = os.getenv("LOGIN_USERNAME", "6301403729")
LOGIN_PASSWORD = os.getenv("LOGIN_PASSWORD", "Dad99128@")

# ============================================================
# CONNECTION POOL — fixes Neon/remote PostgreSQL slowness
#
# IMPORTANT FOR NEON USERS:
#   - If using Neon POOLER url  (contains "-pooler." in hostname):
#     → keepalives and options are NOT supported → removed
#   - If using Neon DIRECT url  (no "-pooler." in hostname):
#     → keepalives work fine
#
# The pool keeps 2 warm connections so every button click
# reuses an existing connection instead of doing a new
# TCP+SSL handshake (which costs 300-800ms each time).
# ============================================================

def _build_pool():
    db_url = os.getenv("DATABASE_URL", "")
    is_pooler = "-pooler." in db_url   # Neon pooler URL detection

    kwargs = dict(
        minconn=1,
        maxconn=10,
        dsn=db_url,
        connect_timeout=10,
    )

    # Neon pooler does NOT support keepalives or options param
    # Direct connections DO support them — keeps idle connections alive
    if not is_pooler:
        kwargs["keepalives"] = 1
        kwargs["keepalives_idle"] = 30
        kwargs["keepalives_interval"] = 5
        kwargs["keepalives_count"] = 3

    return pool.ThreadedConnectionPool(**kwargs)

db_pool = _build_pool()


# ============================================================
# AUTH
# ============================================================

def login_required(f):
    @wraps(f)
    def wrapper(*args, **kwargs):
        if not session.get("logged_in"):
            return redirect("/login")
        return f(*args, **kwargs)
    return wrapper


@app.route("/login", methods=["GET", "POST"])
def login():
    error = ""
    if request.method == "POST":
        username = request.form.get("username")
        password = request.form.get("password")
        if username == LOGIN_USERNAME and password == LOGIN_PASSWORD:
            session["logged_in"] = True
            return redirect("/")
        else:
            error = "Invalid username or password"
    return render_template("login.html", error=error)


@app.route("/logout")
def logout():
    session.clear()
    return redirect("/login")


@app.after_request
def add_header(response):
    response.cache_control.no_store = True
    return response


# ============================================================
# DATABASE
# ============================================================

def connect_db():
    """Get a connection from the pool instead of opening a new one.
    Always call release_conn(conn) when done — this returns it to the pool,
    it does NOT actually close the TCP connection."""
    try:
        conn = db_pool.getconn()
        conn.autocommit = False
        return conn
    except Exception as e:
        print(f"Pool error, falling back: {e}")
        return psycopg2.connect(os.getenv("DATABASE_URL"), connect_timeout=10)

def release_conn(conn):
    try:
        if conn and not conn.closed:
            conn.rollback()
            db_pool.putconn(conn)
    except Exception:
        try:
            conn.close()
        except Exception:
            pass


def create_tables():
    conn = connect_db()
    cursor = conn.cursor()

    cursor.execute("""
        CREATE TABLE IF NOT EXISTS customers (
            id SERIAL PRIMARY KEY,
            name TEXT NOT NULL,
            phone TEXT,
            address TEXT,
            opening_balance INTEGER DEFAULT 0,
            upload_id INTEGER
        )
    """)

    cursor.execute("""
        CREATE TABLE IF NOT EXISTS products (
            id SERIAL PRIMARY KEY,
            name TEXT NOT NULL,
            stock INTEGER DEFAULT 0
        )
    """)

    cursor.execute("""
        CREATE TABLE IF NOT EXISTS sales (
            id SERIAL PRIMARY KEY,
            customer_id INTEGER,
            item TEXT,
            quantity INTEGER,
            total INTEGER DEFAULT 0,
            labour INTEGER DEFAULT 0,
            transport INTEGER DEFAULT 0,
            paid INTEGER DEFAULT 0,
            pending INTEGER DEFAULT 0,
            date TEXT
        )
    """)

    cursor.execute("""
        CREATE TABLE IF NOT EXISTS sale_items (
            id SERIAL PRIMARY KEY,
            sale_id INTEGER,
            product_id INTEGER,
            item TEXT,
            quantity INTEGER,
            price INTEGER,
            total INTEGER
        )
    """)

    cursor.execute("""
        CREATE TABLE IF NOT EXISTS payments (
            id SERIAL PRIMARY KEY,
            customer_id INTEGER,
            amount INTEGER,
            mode TEXT,
            reference TEXT,
            notes TEXT,
            date TEXT
        )
    """)

    cursor.execute("""
        CREATE TABLE IF NOT EXISTS workers (
            id SERIAL PRIMARY KEY,
            name TEXT NOT NULL,
            phone TEXT,
            salary_per_day INTEGER DEFAULT 0
        )
    """)

    cursor.execute("""
        CREATE TABLE IF NOT EXISTS attendance (
            id SERIAL PRIMARY KEY,
            worker_id INTEGER,
            status TEXT,
            check_in TEXT,
            check_out TEXT,
            date TEXT
        )
    """)

    cursor.execute("""
        CREATE TABLE IF NOT EXISTS worker_payments (
            id SERIAL PRIMARY KEY,
            worker_id INTEGER,
            amount INTEGER,
            note TEXT,
            date TEXT
        )
    """)

    cursor.execute("""
        CREATE TABLE IF NOT EXISTS worker_salary_history (
            id SERIAL PRIMARY KEY,
            worker_id INTEGER,
            salary_per_day INTEGER DEFAULT 0,
            from_date TEXT
        )
    """)

    cursor.execute("""
        CREATE TABLE IF NOT EXISTS suppliers (
            id SERIAL PRIMARY KEY,
            name TEXT,
            phone TEXT
        )
    """)

    cursor.execute("""
        CREATE TABLE IF NOT EXISTS purchases (
            id SERIAL PRIMARY KEY,
            supplier_id INTEGER,
            total INTEGER,
            paid INTEGER,
            pending INTEGER,
            date TEXT
        )
    """)

    cursor.execute("""
        CREATE TABLE IF NOT EXISTS supplier_payments (
            id SERIAL PRIMARY KEY,
            supplier_id INTEGER,
            amount INTEGER,
            note TEXT,
            date TEXT
        )
    """)

    cursor.execute("""
        CREATE TABLE IF NOT EXISTS diesel_expenses (
            id SERIAL PRIMARY KEY,
            amount INTEGER,
            note TEXT,
            date TEXT
        )
    """)

    cursor.execute("""
        CREATE TABLE IF NOT EXISTS uploads (
            id SERIAL PRIMARY KEY,
            filename TEXT,
            total_records INTEGER,
            status TEXT,
            date TEXT
        )
    """)

    conn.commit()
    release_conn(conn)

create_tables()


# ============================================================
# SAFE ALTER TABLE (PostgreSQL ignores duplicate column errors
# if we catch them individually per connection)
# ============================================================

def safe_alter(sql):
    try:
        conn = connect_db()
        cursor = conn.cursor()
        cursor.execute(sql)
        conn.commit()
        release_conn(conn)
    except Exception:
        pass

safe_alter("ALTER TABLE worker_salary_history ADD COLUMN salary_per_day INTEGER DEFAULT 0")
safe_alter("ALTER TABLE sales ADD COLUMN labour INTEGER DEFAULT 0")
safe_alter("ALTER TABLE sales ADD COLUMN transport INTEGER DEFAULT 0")
safe_alter("ALTER TABLE customers ADD COLUMN upload_id INTEGER")
safe_alter("ALTER TABLE sale_items ADD COLUMN product_id INTEGER")
safe_alter("ALTER TABLE supplier_payments ADD COLUMN note TEXT DEFAULT ''")


# ============================================================
# DASHBOARD
# ============================================================

@app.route("/")
@login_required
def home():
    conn = connect_db()
    cursor = conn.cursor()

    cursor.execute("SELECT COUNT(*) FROM customers")
    total_customers = cursor.fetchone()[0] or 0

    cursor.execute("SELECT COALESCE(SUM(opening_balance),0) FROM customers")
    opening_total = cursor.fetchone()[0] or 0

    cursor.execute("SELECT COALESCE(SUM(total),0), COALESCE(SUM(paid),0) FROM sales")
    total_sales, sales_paid = cursor.fetchone()

    cursor.execute("SELECT COALESCE(SUM(amount),0) FROM payments")
    payments_total = cursor.fetchone()[0] or 0

    total_collections = sales_paid + payments_total
    total_pending = (opening_total + total_sales) - total_collections

    # FIX: DATE('now') → CURRENT_DATE
    cursor.execute("""
        SELECT COALESCE(SUM(total),0), COUNT(*)
        FROM sales WHERE date = CURRENT_DATE::TEXT
    """)
    today_sales, today_sales_count = cursor.fetchone()

    cursor.execute("""
        SELECT 
        COALESCE((SELECT SUM(amount) FROM payments WHERE date = CURRENT_DATE::TEXT),0)
        +
        COALESCE((SELECT SUM(paid) FROM sales WHERE date = CURRENT_DATE::TEXT),0),

        COALESCE((SELECT COUNT(*) FROM payments WHERE date = CURRENT_DATE::TEXT),0)
        +
        COALESCE((SELECT COUNT(*) FROM sales WHERE date = CURRENT_DATE::TEXT AND paid > 0),0)
    """)
    today_collections, today_collections_count = cursor.fetchone()

    cursor.execute("""
        SELECT COALESCE(SUM(pending),0), COUNT(*)
        FROM sales
        WHERE date = CURRENT_DATE::TEXT AND pending > 0
    """)
    today_pending, today_pending_count = cursor.fetchone()

    # FIX: strftime('%Y-%m', date) → TO_CHAR(CURRENT_DATE, 'YYYY-MM')
    cursor.execute("""
        SELECT COALESCE(SUM(total),0), COUNT(*)
        FROM sales
        WHERE TO_CHAR(date::DATE, 'YYYY-MM') = TO_CHAR(CURRENT_DATE, 'YYYY-MM')
    """)
    monthly_sales, monthly_sales_count = cursor.fetchone()

    cursor.execute("""
        SELECT 
            COALESCE((SELECT SUM(amount) FROM payments WHERE TO_CHAR(date::DATE, 'YYYY-MM') = TO_CHAR(CURRENT_DATE, 'YYYY-MM')),0)
            +
            COALESCE((SELECT SUM(paid) FROM sales WHERE TO_CHAR(date::DATE, 'YYYY-MM') = TO_CHAR(CURRENT_DATE, 'YYYY-MM')),0),

            COALESCE((SELECT COUNT(*) FROM payments WHERE TO_CHAR(date::DATE, 'YYYY-MM') = TO_CHAR(CURRENT_DATE, 'YYYY-MM')),0)
            +
            COALESCE((SELECT COUNT(*) FROM sales WHERE TO_CHAR(date::DATE, 'YYYY-MM') = TO_CHAR(CURRENT_DATE, 'YYYY-MM') AND paid > 0),0)
    """)
    monthly_collections, monthly_collections_count = cursor.fetchone()

    cursor.execute("""
        SELECT COALESCE(SUM(pending),0), COUNT(*)
        FROM sales
        WHERE TO_CHAR(date::DATE, 'YYYY-MM') = TO_CHAR(CURRENT_DATE, 'YYYY-MM') AND pending > 0
    """)
    monthly_pending, monthly_pending_count = cursor.fetchone()

    # FIX: strftime('%d', date) → TO_CHAR(date::DATE, 'DD')
    # FIX: GROUP_CONCAT → STRING_AGG (not needed here, just date aggregation)
    cursor.execute("""
        SELECT TO_CHAR(date::DATE, 'DD'), COALESCE(SUM(total),0)
        FROM sales
        WHERE TO_CHAR(date::DATE, 'YYYY-MM') = TO_CHAR(CURRENT_DATE, 'YYYY-MM')
        GROUP BY date
        ORDER BY date
    """)
    sales_rows = cursor.fetchall()

    cursor.execute("""
        SELECT day, SUM(amount)
        FROM (
            SELECT TO_CHAR(date::DATE, 'DD') AS day, amount
            FROM payments
            WHERE TO_CHAR(date::DATE, 'YYYY-MM') = TO_CHAR(CURRENT_DATE, 'YYYY-MM')

            UNION ALL

            SELECT TO_CHAR(date::DATE, 'DD') AS day, paid AS amount
            FROM sales
            WHERE TO_CHAR(date::DATE, 'YYYY-MM') = TO_CHAR(CURRENT_DATE, 'YYYY-MM')
            AND paid > 0
        ) sub
        GROUP BY day
        ORDER BY day
    """)
    collection_rows = cursor.fetchall()

    sales_labels = [r[0] for r in sales_rows]
    sales_values = [r[1] for r in sales_rows]
    collection_labels = [r[0] for r in collection_rows]
    collection_values = [r[1] for r in collection_rows]

    cursor.execute("""
        SELECT 'Sale', c.name, s.total,
               CASE WHEN s.pending > 0 THEN 'Pending' ELSE 'Paid' END,
               s.date
        FROM sales s
        JOIN customers c ON s.customer_id = c.id

        UNION ALL

        SELECT 'Collection', c.name, p.amount, 'Received', p.date
        FROM payments p
        JOIN customers c ON p.customer_id = c.id

        ORDER BY date DESC
        LIMIT 6
    """)
    recent_transactions = cursor.fetchall()

    release_conn(conn)

    return render_template(
        "index.html",
        total_customers=total_customers,
        total_sales=total_sales,
        total_collections=total_collections,
        total_pending=total_pending,
        today_sales=today_sales,
        today_sales_count=today_sales_count,
        today_collections=today_collections,
        today_collections_count=today_collections_count,
        today_pending=today_pending,
        today_pending_count=today_pending_count,
        monthly_sales=monthly_sales,
        monthly_sales_count=monthly_sales_count,
        monthly_collections=monthly_collections,
        monthly_collections_count=monthly_collections_count,
        monthly_pending=monthly_pending,
        monthly_pending_count=monthly_pending_count,
        sales_labels=sales_labels,
        sales_values=sales_values,
        collection_labels=collection_labels,
        collection_values=collection_values,
        recent_transactions=recent_transactions
    )


# ============================================================
# ADD CUSTOMER
# ============================================================

@app.route("/add_customer", methods=["GET", "POST"])
@login_required
def add_customer():
    if request.method == "POST":
        name = request.form["name"]
        phone = request.form["phone"]
        address = request.form["address"]

        conn = connect_db()
        cursor = conn.cursor()
        cursor.execute(
            "INSERT INTO customers (name, phone, address) VALUES (%s, %s, %s)",
            (name, phone, address)
        )
        conn.commit()
        release_conn(conn)
        return redirect("/")

    return render_template("add_customer.html")


# ============================================================
# SEARCH CUSTOMERS
# FIX: ? → %s, LIKE → ILIKE (case-insensitive in PostgreSQL)
# ============================================================

@app.route("/search", methods=["GET", "POST"])
@login_required
def search():
    conn = connect_db()
    cursor = conn.cursor()

    if request.method == "POST":
        query = request.form.get("query")
        if query:
            cursor.execute("""
                SELECT id, name, phone, address 
                FROM customers
                WHERE name ILIKE %s OR phone ILIKE %s
                ORDER BY id ASC
            """, ('%' + query + '%', '%' + query + '%'))
        else:
            cursor.execute("SELECT id, name, phone, address FROM customers ORDER BY id ASC")
    else:
        cursor.execute("SELECT id, name, phone, address FROM customers ORDER BY id ASC")

    results = cursor.fetchall()
    release_conn(conn)
    return render_template("search.html", results=results)


# ============================================================
# ADD SALE
# FIX: lastrowid → RETURNING id
# FIX: ? → %s
# ============================================================

@app.route("/add_sale", methods=["GET", "POST"])
@login_required
def add_sale():
    conn = connect_db()
    cursor = conn.cursor()

    cursor.execute("SELECT DISTINCT id, name, phone, address FROM customers")
    customers = cursor.fetchall()

    cursor.execute("SELECT id, name, COALESCE(stock,0) FROM products")
    products = cursor.fetchall()

    if request.method == "POST":
        try:
            customer_id = request.form["customer_id"]
            product_ids = request.form.getlist("product_id[]")
            qtys = request.form.getlist("qty[]")
            prices = request.form.getlist("price[]")
            labour = int(request.form.get("labour") or 0)
            transport = int(request.form.get("transport") or 0)
            paid = int(request.form.get("paid") or 0)
            sale_date = request.form.get("sale_date")

            if not sale_date:
                sale_date = datetime.datetime.today().strftime("%Y-%m-%d")

            items_total = 0
            for i in range(len(product_ids)):
                try:
                    q = int(qtys[i]) if qtys[i] else 0
                    p = int(prices[i]) if prices[i] else 0
                    items_total += q * p
                except:
                    pass

            total = items_total + labour + transport
            pending = total - paid

            # FIX: RETURNING id instead of lastrowid
            cursor.execute("""
                INSERT INTO sales (customer_id, total, labour, transport, paid, pending, date)
                VALUES (%s, %s, %s, %s, %s, %s, %s)
                RETURNING id
            """, (customer_id, total, labour, transport, paid, pending, sale_date))
            sale_id = cursor.fetchone()[0]

            for i in range(len(product_ids)):
                try:
                    pid = int(product_ids[i])
                    qty = int(qtys[i]) if qtys[i] else 0
                    price = int(prices[i]) if prices[i] else 0

                    if qty <= 0:
                        continue

                    # FIX: ? → %s
                    cursor.execute("""
                        SELECT name, COALESCE(stock,0)
                        FROM products
                        WHERE id = %s
                    """, (pid,))
                    row = cursor.fetchone()

                    if not row:
                        continue

                    item_name = row[0]
                    available_stock = row[1]

                    if qty > available_stock:
                        conn.rollback()
                        release_conn(conn)
                        return f"""
                        ❌ Not enough stock for {item_name}<br>
                        Available: {available_stock}<br>
                        Trying to sell: {qty}
                        """

                    item_total = qty * price

                    cursor.execute("""
                        INSERT INTO sale_items (sale_id, item, product_id, quantity, price, total)
                        VALUES (%s, %s, %s, %s, %s, %s)
                    """, (sale_id, item_name, pid, qty, price, item_total))

                    cursor.execute("""
                        UPDATE products
                        SET stock = stock - %s
                        WHERE id = %s
                    """, (qty, pid))

                except Exception as e:
                    print("Item Error:", e)

            conn.commit()
            release_conn(conn)
            recalculate_customer_pending(int(customer_id))
            return redirect("/")

        except Exception as e:
            conn.rollback()
            release_conn(conn)
            return f"Error: {str(e)}"

    return render_template("add_sale.html", customers=customers, products=products)


# ============================================================
# UPLOAD CUSTOMERS
# FIX: lastrowid → RETURNING id
# FIX: DATE('now') → CURRENT_DATE::TEXT
# ============================================================

@app.route("/upload_customers", methods=["GET", "POST"])
@login_required
def upload_customers():
    from openpyxl import load_workbook
    from psycopg2.extras import execute_values

    conn = connect_db()
    cursor = conn.cursor()

    try:
        if request.method == "POST":
            file = request.files.get("file")

            if not file or file.filename == "":
                release_conn(conn)
                return redirect("/upload_customers")

            filename = file.filename

            cursor.execute("""
                INSERT INTO uploads (filename, total_records, status, date)
                VALUES (%s, %s, %s, CURRENT_DATE::TEXT)
                RETURNING id
            """, (filename, 0, "Processing"))

            upload_id = cursor.fetchone()[0]

            wb = load_workbook(file, read_only=True, data_only=True)
            ws = wb.active

            rows = ws.iter_rows(values_only=True)

            headers = next(rows)
            headers = [
                str(h).strip().replace("\n", " ").replace("\xa0", " ").upper()
                if h is not None else ""
                for h in headers
            ]

            def col_index(possible_names):
                for name in possible_names:
                    name = name.strip().upper()
                    if name in headers:
                        return headers.index(name)
                return None

            name_i = col_index(["NAME OF THE CONSUMER", "NAME", "CUSTOMER NAME", "CONSUMER NAME"])
            phone_i = col_index(["MOBILE NO", "MOBILE", "PHONE", "PHONE NO", "CONTACT"])
            address_i = col_index(["VILLAGE", "ADDRESS", "PLACE"])
            pending_i = col_index(["PENDING AMOUNT", "PENDING", "OPENING BALANCE", "OLD PENDING", "BALANCE"])

            data_to_insert = []

            for row in rows:
                name = row[name_i] if name_i is not None and name_i < len(row) else ""
                phone = row[phone_i] if phone_i is not None and phone_i < len(row) else ""
                address = row[address_i] if address_i is not None and address_i < len(row) else ""
                opening_balance = row[pending_i] if pending_i is not None and pending_i < len(row) else 0

                if name is None or str(name).strip() == "":
                    continue

                name = str(name).strip()
                phone = "" if phone is None else str(phone).strip()
                address = "" if address is None else str(address).strip()

                try:
                    opening_balance = int(float(str(opening_balance).replace(",", "").replace("₹", "").strip()))
                except:
                    opening_balance = 0

                data_to_insert.append((name, phone, address, opening_balance, upload_id))

            if data_to_insert:
                execute_values(cursor, """
                    INSERT INTO customers (name, phone, address, opening_balance, upload_id)
                    VALUES %s
                """, data_to_insert)

            cursor.execute("""
                UPDATE uploads
                SET total_records=%s, status='Success'
                WHERE id=%s
            """, (len(data_to_insert), upload_id))

            conn.commit()
            release_conn(conn)
            return redirect("/upload_customers")

        cursor.execute("""
            SELECT id, filename, total_records, status, date
            FROM uploads
            ORDER BY id DESC
        """)
        history = cursor.fetchall()

        release_conn(conn)
        return render_template("upload.html", history=history)

    except Exception as e:
        conn.rollback()
        release_conn(conn)
        return f"Upload Error: {str(e)}"

# ============================================================
# DELETE UPLOAD
# FIX: removed sqlite_sequence reset (not applicable in PostgreSQL)
# ============================================================

@app.route("/delete_upload/<int:id>")
@login_required
def delete_upload(id):
    conn = connect_db()
    cursor = conn.cursor()
    cursor.execute("DELETE FROM customers WHERE upload_id=%s", (id,))
    cursor.execute("DELETE FROM uploads WHERE id=%s", (id,))
    conn.commit()
    release_conn(conn)
    return redirect("/upload_customers")


# ============================================================
# CUSTOMER DETAILS
# FIX: ? → %s
# FIX: GROUP_CONCAT → STRING_AGG
# ============================================================

@app.route("/customer/<int:id>")
@login_required
def customer_details(id):
    conn = connect_db()
    cursor = conn.cursor()

    cursor.execute("SELECT * FROM customers WHERE id=%s", (id,))
    customer = cursor.fetchone()

    if not customer:
        release_conn(conn)
        return redirect("/")

    # FIX: GROUP_CONCAT → STRING_AGG
    cursor.execute("""
        SELECT 
            s.id,
            s.total,
            s.paid,
            s.pending,
            s.date,
            STRING_AGG(si.item || ' (' || si.quantity::TEXT || ')', ', '),
            s.labour,
            s.transport
        FROM sales s
        LEFT JOIN sale_items si ON s.id = si.sale_id
        WHERE s.customer_id=%s
        GROUP BY s.id, s.total, s.paid, s.pending, s.date, s.labour, s.transport
        ORDER BY s.id DESC
    """, (id,))
    sales = cursor.fetchall()

    cursor.execute("SELECT SUM(amount) FROM payments WHERE customer_id=%s", (id,))
    payments_total = cursor.fetchone()[0] or 0

    cursor.execute("SELECT SUM(total), SUM(paid), SUM(pending) FROM sales WHERE customer_id=%s", (id,))
    totals = cursor.fetchone()
    sales_total = totals[0] or 0
    sales_paid = totals[1] or 0

    opening = customer[4] or 0
    total = sales_total + opening
    paid = sales_paid + payments_total
    pending = total - paid

    cursor.execute("""
        SELECT amount, mode, reference, notes, date
        FROM payments WHERE customer_id=%s ORDER BY id DESC
    """, (id,))
    payments = cursor.fetchall()
    release_conn(conn)

    return render_template(
        "customer_details.html",
        customer=customer,
        sales=sales,
        total=total,
        paid=paid,
        pending=pending,
        payments=payments
    )


# ============================================================
# VIEW ALL CUSTOMERS
# FIX: ? → %s, LIKE → ILIKE
# ============================================================

@app.route('/customers')
@login_required
def customers():
    conn = connect_db()
    cursor = conn.cursor()
    search = request.args.get("search")

    if search:
        cursor.execute("""
            SELECT id, name, phone, address, opening_balance
            FROM customers
            WHERE name ILIKE %s OR phone ILIKE %s
            ORDER BY id ASC
        """, ('%' + search + '%', '%' + search + '%'))
    else:
        cursor.execute("SELECT id, name, phone, address, opening_balance FROM customers ORDER BY id ASC")

    data = cursor.fetchall()
    release_conn(conn)
    return render_template("customers.html", customers=data)


# ============================================================
# DELETE CUSTOMER
# FIX: ? → %s
# ============================================================

@app.route("/delete_customer/<int:id>")
@login_required
def delete_customer(id):
    conn = connect_db()
    cursor = conn.cursor()
    cursor.execute("DELETE FROM sale_items WHERE sale_id IN (SELECT id FROM sales WHERE customer_id=%s)", (id,))
    cursor.execute("DELETE FROM sales WHERE customer_id=%s", (id,))
    cursor.execute("DELETE FROM payments WHERE customer_id=%s", (id,))
    cursor.execute("DELETE FROM customers WHERE id=%s", (id,))
    conn.commit()
    release_conn(conn)
    return redirect("/customers")


# ============================================================
# PAYMENTS
# FIX: ? → %s, DATE('now') → CURRENT_DATE::TEXT
# FIX: strftime → TO_CHAR
# ============================================================

@app.route("/payments", methods=["GET", "POST"])
@login_required
def payments():
    conn = connect_db()
    cursor = conn.cursor()
    selected_customer_id = request.args.get("customer_id")

    cursor.execute("SELECT id, name, phone, address FROM customers")
    customers = cursor.fetchall()

    if request.method == "POST":
        customer_id = request.form.get("customer_id")
        if not customer_id:
            release_conn(conn)
            return redirect("/payments")

        amount = int(request.form["amount"])
        mode = request.form["mode"]
        reference = request.form.get("reference", "")
        notes = request.form.get("notes", "")

        # FIX: DATE('now') → CURRENT_DATE::TEXT
        cursor.execute("""
            INSERT INTO payments (customer_id, amount, mode, reference, notes, date)
            VALUES (%s, %s, %s, %s, %s, CURRENT_DATE::TEXT)
        """, (customer_id, amount, mode, reference, notes))

        conn.commit()
        release_conn(conn)
        recalculate_customer_pending(customer_id)
        return redirect(f"/payments?customer_id={customer_id}")

    total_purchase = 0
    total_paid = 0
    total_pending = 0
    last_payment = 0
    last_payment_date = ""
    selected_customer = None

    if selected_customer_id:
        cursor.execute("SELECT id, name, phone, address FROM customers WHERE id=%s", (selected_customer_id,))
        selected_customer = cursor.fetchone()

        cursor.execute("SELECT SUM(total), SUM(paid), SUM(pending) FROM sales WHERE customer_id=%s", (selected_customer_id,))
        sales_data = cursor.fetchone()
        sales_total = sales_data[0] or 0
        sales_paid = sales_data[1] or 0

        cursor.execute("SELECT opening_balance FROM customers WHERE id=%s", (selected_customer_id,))
        row = cursor.fetchone()
        opening = row[0] if row else 0

        total_purchase = sales_total + opening

        cursor.execute("SELECT SUM(amount) FROM payments WHERE customer_id=%s", (selected_customer_id,))
        payments_total = cursor.fetchone()[0] or 0

        total_paid = sales_paid + payments_total
        total_pending = total_purchase - total_paid

        cursor.execute("""
            SELECT amount, date FROM payments WHERE customer_id=%s ORDER BY id DESC LIMIT 1
        """, (selected_customer_id,))
        last = cursor.fetchone()
        if last:
            last_payment = last[0]
            last_payment_date = last[1]

    cursor.execute("""
        SELECT p.id, c.name, p.amount, p.mode, p.reference, p.date
        FROM payments p
        JOIN customers c ON p.customer_id = c.id

        UNION ALL

        SELECT s.id, c.name, s.paid, 'Sale Payment', '', s.date
        FROM sales s
        JOIN customers c ON s.customer_id = c.id
        WHERE s.paid > 0

        ORDER BY date DESC
        LIMIT 10
    """)
    payments_list = cursor.fetchall()

    # FIX: DATE('now') → CURRENT_DATE::TEXT
    cursor.execute("""
        SELECT COALESCE((SELECT SUM(amount) FROM payments WHERE date = CURRENT_DATE::TEXT),0)
        + COALESCE((SELECT SUM(paid) FROM sales WHERE date = CURRENT_DATE::TEXT),0)
    """)
    today = cursor.fetchone()[0] or 0

    cursor.execute("""
        SELECT COALESCE((SELECT SUM(amount) FROM payments),0)
        + COALESCE((SELECT SUM(paid) FROM sales WHERE paid > 0),0)
    """)
    total = cursor.fetchone()[0] or 0

    cursor.execute("SELECT SUM(pending) FROM sales")
    sales_pending_all = cursor.fetchone()[0] or 0

    cursor.execute("SELECT SUM(opening_balance) FROM customers")
    opening_all = cursor.fetchone()[0] or 0

    cursor.execute("SELECT SUM(amount) FROM payments")
    payments_all = cursor.fetchone()[0] or 0

    overall_pending = (sales_pending_all + opening_all) - payments_all

    release_conn(conn)

    return render_template(
        "payments.html",
        customers=customers,
        payments=payments_list,
        today=today,
        total=total,
        total_purchase=total_purchase,
        total_paid=total_paid,
        total_pending=total_pending,
        last_payment=last_payment,
        last_payment_date=last_payment_date,
        selected_customer_id=selected_customer_id,
        selected_customer=selected_customer,
        overall_pending=overall_pending
    )


# ============================================================
# DOWNLOAD REPORT
# FIX: ? → %s in queries
# FIX: pd.read_sql_query with psycopg2 — use cursor + DataFrame
# ============================================================

@app.route("/download_report", methods=["GET", "POST"])
@login_required
def download_report():
    if request.method == "POST":
        report_type = request.form.get("type")
        from_date = request.form.get("from_date")
        to_date = request.form.get("to_date")

        conn = connect_db()
        cursor = conn.cursor()

        if report_type == "sales":
            # FIX: GROUP_CONCAT → STRING_AGG, ? → %s
            cursor.execute("""
                SELECT 
                    s.id AS id,
                    c.name AS name,
                    COALESCE(STRING_AGG(si.item || '(' || si.quantity::TEXT || ')', ', '), '') AS item,
                    COALESCE(STRING_AGG(si.quantity::TEXT, ', '), '') AS qtys,
                    s.total AS total,
                    s.paid AS paid,
                    s.pending AS pending,
                    s.date AS date
                FROM sales s
                JOIN customers c ON s.customer_id = c.id
                LEFT JOIN sale_items si ON s.id = si.sale_id
                WHERE s.date BETWEEN %s AND %s
                GROUP BY s.id, c.name, s.total, s.paid, s.pending, s.date
                ORDER BY s.id DESC
            """, (from_date, to_date))
            columns = ["id", "name", "item", "qtys", "total", "paid", "pending", "date"]

        elif report_type == "payments":
            cursor.execute("""
                SELECT 
                    p.id AS id,
                    c.name AS name,
                    p.amount AS amount,
                    p.mode AS mode,
                    p.reference AS reference,
                    p.notes AS notes,
                    p.date AS date
                FROM payments p
                JOIN customers c ON p.customer_id = c.id
                WHERE p.date BETWEEN %s AND %s
                ORDER BY p.id DESC
            """, (from_date, to_date))
            columns = ["id", "name", "amount", "mode", "reference", "notes", "date"]

        else:
            release_conn(conn)
            return "Invalid report type"

        rows = cursor.fetchall()
        release_conn(conn)

        df = pd.DataFrame(rows, columns=columns)

        output = io.BytesIO()
        df.to_excel(output, index=False)
        output.seek(0)

        return send_file(
            output,
            download_name=f"{report_type}_report_{from_date}_to_{to_date}.xlsx",
            as_attachment=True
        )

    return render_template("download.html")


# ============================================================
# EDIT CUSTOMER
# FIX: ? → %s
# ============================================================

@app.route("/edit_customer/<int:id>", methods=["GET", "POST"])
@login_required
def edit_customer(id):
    conn = connect_db()
    cursor = conn.cursor()

    if request.method == "POST":
        name = request.form["name"]
        phone = request.form["phone"]
        address = request.form["address"]

        cursor.execute("""
            UPDATE customers SET name=%s, phone=%s, address=%s WHERE id=%s
        """, (name, phone, address, id))
        conn.commit()
        release_conn(conn)
        return redirect("/customers")

    cursor.execute("SELECT * FROM customers WHERE id=%s", (id,))
    customer = cursor.fetchone()
    release_conn(conn)
    return render_template("edit_customer.html", customer=customer)


# ============================================================
# EDIT SALE
# FIX: ? → %s throughout
# ============================================================

@app.route("/edit_sale/<int:sale_id>", methods=["GET", "POST"])
@login_required
def edit_sale(sale_id):
    conn = connect_db()
    cursor = conn.cursor()

    cursor.execute("SELECT customer_id FROM sales WHERE id=%s", (sale_id,))
    result = cursor.fetchone()

    if not result:
        release_conn(conn)
        return "Sale not found"

    customer_id = result[0]

    if request.method == "POST":

        # Restore old stock
        cursor.execute("SELECT item, quantity FROM sale_items WHERE sale_id=%s", (sale_id,))
        old_items = cursor.fetchall()

        for item, qty in old_items:
            cursor.execute("SELECT id FROM products WHERE name=%s", (item,))
            product = cursor.fetchone()
            if product:
                cursor.execute("UPDATE products SET stock = stock + %s WHERE id = %s", (qty, product[0]))

        items = request.form.getlist("item[]")
        qtys = request.form.getlist("qty[]")
        prices = request.form.getlist("price[]")
        labour = int(request.form.get("labour") or 0)
        transport = int(request.form.get("transport") or 0)
        paid = int(request.form.get("paid") or 0)

        items_total = 0
        for i in range(len(items)):
            try:
                q = int(qtys[i]) if qtys[i] else 0
                p = int(prices[i]) if prices[i] else 0
                items_total += q * p
            except:
                pass

        total = items_total + labour + transport
        pending = total - paid

        cursor.execute("""
            UPDATE sales SET total=%s, labour=%s, transport=%s, paid=%s, pending=%s
            WHERE id=%s
        """, (total, labour, transport, paid, pending, sale_id))

        cursor.execute("DELETE FROM sale_items WHERE sale_id=%s", (sale_id,))

        for i in range(len(items)):
            try:
                item = items[i].strip()
                q = int(qtys[i]) if qtys[i] else 0
                p = int(prices[i]) if prices[i] else 0
                t = q * p

                if item:
                    cursor.execute("""
                        INSERT INTO sale_items (sale_id, item, quantity, price, total)
                        VALUES (%s, %s, %s, %s, %s)
                    """, (sale_id, item, q, p, t))

                    cursor.execute("SELECT id FROM products WHERE name=%s", (item,))
                    product = cursor.fetchone()
                    if product:
                        cursor.execute("UPDATE products SET stock = stock - %s WHERE id = %s", (q, product[0]))
            except:
                pass

        conn.commit()
        release_conn(conn)
        recalculate_customer_pending(customer_id)
        return redirect(f"/customer/{customer_id}")

    cursor.execute("SELECT customer_id, total, labour, transport, paid FROM sales WHERE id=%s", (sale_id,))
    sale = cursor.fetchone()

    cursor.execute("SELECT item, quantity, price FROM sale_items WHERE sale_id=%s", (sale_id,))
    items = cursor.fetchall()
    release_conn(conn)

    return render_template("edit_sale.html", sale=sale, items=items, sale_id=sale_id)


# ============================================================
# EDIT PAYMENT
# FIX: ? → %s
# ============================================================

@app.route("/edit_payment/<int:id>", methods=["GET", "POST"])
@login_required
def edit_payment(id):
    conn = connect_db()
    cursor = conn.cursor()

    if request.method == "POST":
        amount = int(request.form["amount"])
        mode = request.form["mode"]
        reference = request.form.get("reference", "")
        notes = request.form.get("notes", "")

        cursor.execute("SELECT customer_id FROM payments WHERE id=%s", (id,))
        customer_id = cursor.fetchone()[0]

        cursor.execute("""
            UPDATE payments SET amount=%s, mode=%s, reference=%s, notes=%s WHERE id=%s
        """, (amount, mode, reference, notes, id))

        conn.commit()
        release_conn(conn)
        recalculate_customer_pending(customer_id)
        return redirect("/payments")

    cursor.execute("SELECT customer_id FROM payments WHERE id=%s", (id,))
    customer_id = cursor.fetchone()[0]

    cursor.execute("""
        SELECT p.*, c.name FROM payments p JOIN customers c ON p.customer_id = c.id WHERE p.id=%s
    """, (id,))
    payment = cursor.fetchone()
    release_conn(conn)

    recalculate_customer_pending(customer_id)
    return render_template("edit_payment.html", payment=payment)


# ============================================================
# ADD WORKER
# FIX: ? → %s, lastrowid → RETURNING id, DATE('now') → CURRENT_DATE::TEXT
# ============================================================

@app.route("/add_worker", methods=["GET", "POST"])
@login_required
def add_worker():
    if request.method == "POST":
        name = request.form["name"]
        phone = request.form["phone"]
        salary = request.form["salary"]

        conn = connect_db()
        cursor = conn.cursor()

        # FIX: RETURNING id
        cursor.execute(
            "INSERT INTO workers (name, phone, salary_per_day) VALUES (%s, %s, %s) RETURNING id",
            (name, phone, salary)
        )
        worker_id = cursor.fetchone()[0]

        # FIX: DATE('now') → CURRENT_DATE::TEXT
        cursor.execute("""
            INSERT INTO worker_salary_history (worker_id, salary_per_day, from_date)
            VALUES (%s, %s, TO_CHAR(CURRENT_DATE, 'YYYY-MM') || '-01')
        """, (worker_id, salary))

        conn.commit()
        release_conn(conn)
        return redirect("/attendance")

    return render_template("add_worker.html")


# ============================================================
# ATTENDANCE
# ============================================================

@app.route("/attendance")
@login_required
def attendance():
    conn = connect_db()
    try:
        cursor = conn.cursor()

        selected_date = request.args.get("date")
        if not selected_date:
            selected_date = datetime.datetime.today().strftime("%Y-%m-%d")

        filter_type = request.args.get("filter", "all")

        cursor.execute("SELECT COUNT(*) FROM workers")
        total_workers = cursor.fetchone()[0]

        cursor.execute("SELECT COUNT(*) FROM attendance WHERE status='Present' AND date=%s", (selected_date,))
        present = cursor.fetchone()[0]

        cursor.execute("SELECT COUNT(*) FROM attendance WHERE status='Absent' AND date=%s", (selected_date,))
        absent = cursor.fetchone()[0]

        cursor.execute("SELECT COUNT(*) FROM attendance WHERE status='Half' AND date=%s", (selected_date,))
        half = cursor.fetchone()[0]

        query = """
            SELECT a.id, w.name, a.check_in, a.check_out, a.status, a.date
            FROM attendance a
            JOIN workers w ON a.worker_id = w.id
            WHERE a.date=%s
        """
        params = [selected_date]

        if filter_type == "present":
            query += " AND a.status='Present'"
        elif filter_type == "absent":
            query += " AND a.status='Absent'"
        elif filter_type == "half":
            query += " AND a.status='Half'"

        cursor.execute(query, params)
        data = cursor.fetchall()
        
        return render_template("attendance.html",
            data=data,
            total_workers=total_workers,
            present=present,
            absent=absent,
            half=half,
            filter_type=filter_type,
            selected_date=selected_date
        )
    finally:
        release_conn(conn)


# ============================================================
# MARK ATTENDANCE SAVE
# ============================================================

@app.route("/mark_attendance_save", methods=["POST"])
@login_required
def mark_attendance_save():
    conn = connect_db()
    try:
        cursor = conn.cursor()

        date = request.form.get("date")
        if not date:
            date = datetime.datetime.today().strftime("%Y-%m-%d")

        worker_ids = request.form.getlist("worker_ids")

        for wid in worker_ids:
            status = request.form.get(f"status_{wid}")
            check_in = request.form.get(f"check_in_{wid}")
            check_out = request.form.get(f"check_out_{wid}")

            if not status:
                continue

            cursor.execute("DELETE FROM attendance WHERE worker_id=%s AND date=%s", (wid, date))

            cursor.execute("""
                INSERT INTO attendance (worker_id, status, check_in, check_out, date)
                VALUES (%s, %s, %s, %s, %s)
            """, (wid, status, check_in, check_out, date))

        conn.commit()
        return redirect(f"/attendance?date={date}")
    except Exception as e:
        conn.rollback()
        raise e
    finally:
        release_conn(conn)


# ============================================================
# EDIT ATTENDANCE
# ============================================================

@app.route("/edit_attendance/<int:id>", methods=["GET", "POST"])
@login_required
def edit_attendance(id):
    conn = connect_db()
    try:
        cursor = conn.cursor()

        if request.method == "POST":
            status = request.form["status"]
            cursor.execute("UPDATE attendance SET status=%s WHERE id=%s", (status, id))
            conn.commit()
            return redirect("/attendance")

        cursor.execute("SELECT status FROM attendance WHERE id=%s", (id,))
        data = cursor.fetchone()
        return render_template("edit_attendance.html", data=data, id=id)
    finally:
        release_conn(conn)


# ============================================================
# MARK ATTENDANCE PAGE
# ============================================================

@app.route("/mark_attendance")
@login_required
def mark_attendance_page():
    conn = connect_db()
    try:
        cursor = conn.cursor()
        cursor.execute("SELECT id, name FROM workers")
        workers = cursor.fetchall()
        return render_template("mark_attendance.html", workers=workers)
    finally:
        release_conn(conn)


# ============================================================
# ATTENDANCE SUMMARY
# ============================================================

@app.route("/attendance_summary")
@login_required
def attendance_summary():
    conn = connect_db()
    try:
        cursor = conn.cursor()

        month = request.args.get("month")
        if not month:
            month = datetime.datetime.today().strftime("%Y-%m")

        cursor.execute("SELECT id, name, COALESCE(salary_per_day, 0) FROM workers")
        workers = cursor.fetchall()

        final_data = []

        for wid, name, current_salary in workers:
            cursor.execute("""
                SELECT date, status
                FROM attendance
                WHERE worker_id=%s AND TO_CHAR(date::DATE, 'YYYY-MM')=%s
                ORDER BY date ASC
            """, (wid, month))
            records = cursor.fetchall()

            present = 0
            half = 0
            absent = 0
            total_salary = 0

            for date, status in records:
                cursor.execute("""
                    SELECT salary_per_day
                    FROM worker_salary_history
                    WHERE worker_id=%s AND from_date <= %s
                    ORDER BY from_date DESC
                    LIMIT 1
                """, (wid, date))
                row = cursor.fetchone()
                salary = row[0] if (row and row[0]) else (current_salary or 0)

                if status == "Present":
                    present += 1
                    total_salary += salary
                elif status == "Half":
                    half += 1
                    total_salary += salary * 0.5
                elif status == "Absent":
                    absent += 1

            total_days = present + (half * 0.5)

            cursor.execute("""
                SELECT COALESCE(SUM(amount), 0)
                FROM worker_payments
                WHERE worker_id=%s AND TO_CHAR(date::DATE, 'YYYY-MM')=%s
            """, (wid, month))
            advance = cursor.fetchone()[0] or 0

            final_salary = total_salary - advance

            final_data.append({
                "id": wid,
                "name": name,
                "present": present,
                "half": half,
                "absent": absent,
                "total_days": total_days,
                "total_salary": total_salary,
                "advance": advance,
                "salary": final_salary
            })

        return render_template("attendance_summary.html", data=final_data, month=month)
    finally:
        release_conn(conn)


# ============================================================
# UPDATE SALARY
# ============================================================

@app.route("/update_salary", methods=["GET", "POST"])
@login_required
def update_salary():
    conn = connect_db()
    try:
        cursor = conn.cursor()

        cursor.execute("""
            SELECT id, name, salary_per_day
            FROM workers
            ORDER BY name ASC
        """)
        workers = cursor.fetchall()

        if request.method == "POST":
            worker_id = request.form["worker_id"]
            new_salary = int(request.form["salary"])
            from_month = request.form["from_month"]
            from_date = from_month + "-01"

            # ✅ Update latest/current salary shown in UI
            cursor.execute("""
                UPDATE workers
                SET salary_per_day=%s
                WHERE id=%s
            """, (new_salary, worker_id))

            # ✅ Save salary history for correct month-wise calculation
            cursor.execute("""
                INSERT INTO worker_salary_history (worker_id, salary_per_day, from_date)
                VALUES (%s, %s, %s)
            """, (worker_id, new_salary, from_date))

            conn.commit()
            return redirect("/attendance_summary")

        return render_template("update_salary.html", workers=workers)

    except Exception as e:
        conn.rollback()
        raise e

    finally:
        release_conn(conn)


# ============================================================
# WORKER PAYMENTS
# ============================================================

@app.route("/worker_payments", methods=["GET", "POST"])
@login_required
def worker_payments():
    conn = connect_db()
    try:
        cursor = conn.cursor()

        cursor.execute("SELECT id, name FROM workers")
        workers = cursor.fetchall()

        if request.method == "POST":
            worker_id = request.form["worker_id"]
            amount = int(request.form["amount"])
            note = request.form.get("note", "")

            # CURRENT_DATE is universally understood by Postgres natively
            cursor.execute("""
                INSERT INTO worker_payments (worker_id, amount, note, date)
                VALUES (%s, %s, %s, CURRENT_DATE)
            """, (worker_id, amount, note))

            conn.commit()
            return redirect("/worker_payments")

        cursor.execute("""
            SELECT w.name, p.amount, p.note, p.date
            FROM worker_payments p
            JOIN workers w ON p.worker_id = w.id
            ORDER BY p.id DESC
        """)
        payments = cursor.fetchall()
        return render_template("worker_payments.html", workers=workers, payments=payments)
    except Exception as e:
        conn.rollback()
        raise e
    finally:
        release_conn(conn)




# ============================================================
# DOWNLOAD SALARY REPORT (PDF)
# FIX: ? → %s, strftime → TO_CHAR
# ============================================================

from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape

@app.route("/download_salary_report")
@login_required
def download_salary_report():
    conn = connect_db()
    cursor = conn.cursor()

    month = request.args.get("month")
    if not month:
        return "❌ Please select a month first!"

    month_name = datetime.datetime.strptime(month, "%Y-%m").strftime("%B %Y")

    cursor.execute("SELECT id, name FROM workers")
    workers = cursor.fetchall()

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer)
    styles = getSampleStyleSheet()
    elements = []

    elements.append(Paragraph(f"<b>Monthly Salary Report - {month_name}</b>", styles['Title']))
    elements.append(Spacer(1, 20))

    for wid, name in workers:
        # FIX: strftime → TO_CHAR
        cursor.execute("""
            SELECT date, status
            FROM attendance
            WHERE worker_id=%s AND TO_CHAR(date::DATE, 'YYYY-MM')=%s
        """, (wid, month))
        records = cursor.fetchall()

        present = 0
        half = 0
        absent = 0
        total_salary = 0

        for date, status in records:
            cursor.execute("""
                SELECT salary_per_day
                FROM worker_salary_history
                WHERE worker_id=%s AND from_date <= %s
                ORDER BY from_date DESC LIMIT 1
            """, (wid, date))
            row = cursor.fetchone()
            salary = row[0] if row else 0

            if status == "Present":
                present += 1
                total_salary += salary
            elif status == "Half":
                half += 1
                total_salary += salary * 0.5
            elif status == "Absent":
                absent += 1

        cursor.execute("""
            SELECT amount, date FROM worker_payments
            WHERE worker_id=%s AND TO_CHAR(date::DATE, 'YYYY-MM')=%s
        """, (wid, month))
        advances = cursor.fetchall()

        advance_total = sum([a[0] for a in advances]) if advances else 0
        final_salary = total_salary - advance_total

        elements.append(Paragraph(f"<b>Worker:</b> {name}", styles['Heading3']))
        elements.append(Paragraph(f"Present: {present} | Half: {half} | Absent: {absent}", styles['Normal']))
        elements.append(Paragraph(f"<b>Total Salary:</b> ₹{total_salary}", styles['Normal']))
        elements.append(Spacer(1, 8))
        elements.append(Paragraph("<b>Advances:</b>", styles['Normal']))

        if advances:
            for amt, d in advances:
                elements.append(Paragraph(f"₹{amt} - {d}", styles['Normal']))
        else:
            elements.append(Paragraph("No advances", styles['Normal']))

        elements.append(Spacer(1, 8))
        elements.append(Paragraph(f"<b>Total Advance:</b> ₹{advance_total}", styles['Normal']))
        elements.append(Paragraph(f"<b>Final Salary:</b> ₹{final_salary}", styles['Normal']))
        elements.append(Spacer(1, 12))
        elements.append(Paragraph("-" * 60, styles['Normal']))
        elements.append(Spacer(1, 12))

    release_conn(conn)
    doc.build(elements)
    buffer.seek(0)

    return send_file(buffer, as_attachment=True, download_name=f"salary_report_{month}.pdf", mimetype='application/pdf')


# ============================================================
# SUPPLIERS
# FIX: ? → %s, DATE('now') → CURRENT_DATE::TEXT
# ============================================================

@app.route("/suppliers", methods=["GET", "POST"])
@login_required
def suppliers():
    conn = connect_db()
    cursor = conn.cursor()

    if request.method == "POST" and request.form.get("form_type") == "add_supplier":
        name = request.form["name"]
        phone = request.form["phone"]
        cursor.execute("INSERT INTO suppliers (name, phone) VALUES (%s, %s)", (name, phone))
        conn.commit()
        release_conn(conn)
        return redirect("/suppliers")

    if request.method == "POST" and request.form.get("form_type") == "add_purchase":
        supplier_id = request.form["supplier_id"]
        total = int(request.form["total"])
        paid = int(request.form["paid"])
        cursor.execute("""
            INSERT INTO purchases (supplier_id, total, paid, pending, date)
            VALUES (%s, %s, %s, %s, CURRENT_DATE::TEXT)
        """, (supplier_id, total, paid, total - paid))
        conn.commit()
        release_conn(conn)
        recalculate_supplier_pending(supplier_id)
        return redirect("/suppliers")

    if request.method == "POST" and request.form.get("form_type") == "add_payment":
        supplier_id = request.form["supplier_id"]
        amount = int(request.form["amount"])
        note = request.form.get("note", "")
        cursor.execute("""
            INSERT INTO supplier_payments (supplier_id, amount, note, date)
            VALUES (%s, %s, %s, CURRENT_DATE::TEXT)
        """, (supplier_id, amount, note))
        conn.commit()
        release_conn(conn)
        recalculate_supplier_pending(supplier_id)
        return redirect("/suppliers")

    cursor.execute("SELECT * FROM suppliers")
    suppliers_list = cursor.fetchall()

    cursor.execute("""
        SELECT p.id, s.name, p.total, p.paid, p.pending, p.date
        FROM purchases p JOIN suppliers s ON p.supplier_id = s.id ORDER BY p.id DESC
    """)
    purchases = cursor.fetchall()

    cursor.execute("""
        SELECT sp.id, s.name, sp.amount, sp.date, COALESCE(sp.note, '')
        FROM supplier_payments sp JOIN suppliers s ON sp.supplier_id = s.id ORDER BY sp.id DESC
    """)
    payments = cursor.fetchall()

    cursor.execute("""
        SELECT 
            s.name,
            COALESCE(SUM(p.total), 0),
            COALESCE(SUM(p.paid), 0),
            COALESCE((SELECT SUM(sp.amount) FROM supplier_payments sp WHERE sp.supplier_id = s.id), 0),
            COALESCE(SUM(p.paid), 0) + COALESCE((SELECT SUM(sp.amount) FROM supplier_payments sp WHERE sp.supplier_id = s.id), 0),
            COALESCE(SUM(p.pending), 0)
        FROM suppliers s
        LEFT JOIN purchases p ON s.id = p.supplier_id
        GROUP BY s.id, s.name
    """)
    summary = cursor.fetchall()

    release_conn(conn)
    return render_template("suppliers.html", suppliers=suppliers_list, purchases=purchases, payments=payments, summary=summary)


# ============================================================
# EDIT PURCHASE
# FIX: ? → %s
# ============================================================

@app.route("/edit_purchase/<int:id>", methods=["GET", "POST"])
@login_required
def edit_purchase(id):
    conn = connect_db()
    cursor = conn.cursor()

    if request.method == "POST":
        total = int(float(request.form["total"].strip()))
        paid = int(float(request.form["paid"].strip()))
        pending = total - paid

        cursor.execute("UPDATE purchases SET total=%s, paid=%s, pending=%s WHERE id=%s", (total, paid, pending, id))
        conn.commit()
        release_conn(conn)
        return redirect("/suppliers")

    cursor.execute("SELECT id, total, paid FROM purchases WHERE id=%s", (id,))
    purchase = cursor.fetchone()
    release_conn(conn)
    return render_template("edit_purchase.html", p=purchase)


# ============================================================
# DELETE PURCHASE
# FIX: ? → %s
# ============================================================

@app.route("/delete_purchase/<int:id>")
@login_required
def delete_purchase(id):
    conn = connect_db()
    cursor = conn.cursor()
    cursor.execute("DELETE FROM purchases WHERE id=%s", (id,))
    conn.commit()
    release_conn(conn)
    return redirect("/suppliers")


# ============================================================
# DELETE SUPPLIER PAYMENT
# FIX: ? → %s
# ============================================================

@app.route("/delete_supplier_payment/<int:id>")
@login_required
def delete_supplier_payment(id):
    conn = connect_db()
    cursor = conn.cursor()
    cursor.execute("SELECT supplier_id FROM supplier_payments WHERE id=%s", (id,))
    row = cursor.fetchone()

    if row:
        supplier_id = row[0]
        cursor.execute("DELETE FROM supplier_payments WHERE id=%s", (id,))
        conn.commit()
        release_conn(conn)
        recalculate_supplier_pending(supplier_id)

    return redirect("/suppliers")


# ============================================================
# EDIT SUPPLIER PAYMENT
# FIX: ? → %s
# ============================================================

@app.route("/edit_supplier_payment/<int:id>", methods=["GET", "POST"])
@login_required
def edit_supplier_payment(id):
    conn = connect_db()
    cursor = conn.cursor()

    if request.method == "POST":
        amount = int(float(request.form["amount"].strip()))
        cursor.execute("SELECT supplier_id FROM supplier_payments WHERE id=%s", (id,))
        supplier_id = cursor.fetchone()[0]
        cursor.execute("UPDATE supplier_payments SET amount=%s WHERE id=%s", (amount, id))
        conn.commit()
        release_conn(conn)
        recalculate_supplier_pending(supplier_id)
        return redirect("/suppliers")

    cursor.execute("SELECT id, amount FROM supplier_payments WHERE id=%s", (id,))
    payment = cursor.fetchone()
    release_conn(conn)
    return render_template("edit_supplier_payment.html", p=payment)


# ============================================================
# INVOICE
# FIX: ? → %s
# ============================================================

@app.route("/invoice/<int:sale_id>")
@login_required
def invoice(sale_id):
    conn = connect_db()
    cursor = conn.cursor()

    cursor.execute("""
        SELECT customer_id, total, labour, transport, paid, pending, date
        FROM sales WHERE id=%s
    """, (sale_id,))
    s = cursor.fetchone()

    cursor.execute("SELECT name, phone, address FROM customers WHERE id=%s", (s[0],))
    c = cursor.fetchone()

    cursor.execute("SELECT item, quantity, price FROM sale_items WHERE sale_id=%s", (sale_id,))
    items = cursor.fetchall()
    release_conn(conn)

    customer = ("", c[0], c[1], c[2])
    sale = {
        "labour": s[2],
        "transport": s[3],
        "paid": s[4],
        "pending": s[5],
        "date": s[6]
    }

    file_path = f"invoice_{sale_id}.pdf"
    generate_invoice(file_path, customer, items, sale)
    return send_file(file_path, as_attachment=True)


# ============================================================
# INVOICE GENERATION
# ============================================================

def generate_invoice(file_path, customer, items, sale):
    doc = SimpleDocTemplate(file_path)
    styles = getSampleStyleSheet()
    elements = []

    elements.append(Paragraph("<b>CEMENT STORE</b>", styles['Title']))
    elements.append(Spacer(1, 10))
    elements.append(Paragraph(f"Customer: {customer[1]}", styles['Normal']))
    elements.append(Paragraph(f"Phone: {customer[2]}", styles['Normal']))
    elements.append(Paragraph(f"Address: {customer[3]}", styles['Normal']))
    elements.append(Paragraph(f"Date: {sale['date']}", styles['Normal']))
    elements.append(Spacer(1, 15))

    data = [["Item", "Qty", "Price", "Amount"]]
    items_total = 0

    for i in items:
        item_name = i[0]
        qty = i[1]
        price = i[2]
        amount = qty * price
        items_total += amount
        data.append([item_name, str(qty), f"₹{price}", f"₹{amount}"])

    table = Table(data)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("GRID", (0, 0), (-1, -1), 1, colors.black),
        ("ALIGN", (1, 1), (-1, -1), "CENTER")
    ]))
    elements.append(table)
    elements.append(Spacer(1, 20))

    labour = sale["labour"]
    transport = sale["transport"]
    grand_total = items_total + labour + transport

    elements.append(Paragraph(f"Labour: ₹{labour}", styles['Normal']))
    elements.append(Paragraph(f"Transport: ₹{transport}", styles['Normal']))
    elements.append(Spacer(1, 10))
    elements.append(Paragraph(f"<b>Total: ₹{grand_total}</b>", styles['Normal']))
    elements.append(Paragraph(f"Paid: ₹{sale['paid']}", styles['Normal']))
    elements.append(Paragraph(f"Pending: ₹{sale['pending']}", styles['Normal']))
    elements.append(Spacer(1, 20))
    elements.append(Paragraph("Thank you for your business!", styles['Italic']))

    doc.build(elements)


# ============================================================
# STOCK MANAGEMENT
# FIX: ? → %s
# ============================================================

@app.route("/stock", methods=["GET", "POST"])
@login_required
def stock():
    conn = connect_db()
    cursor = conn.cursor()

    if request.method == "POST" and request.form.get("form_type") == "add_stock":
        product_id = request.form["product_id"]
        added = int(request.form.get("added", 0))
        cursor.execute("UPDATE products SET stock = COALESCE(stock, 0) + %s WHERE id = %s", (added, product_id))
        conn.commit()
        release_conn(conn)
        return redirect("/stock")

    if request.method == "POST" and request.form.get("form_type") == "add_product":
        name = request.form["name"]
        cursor.execute("INSERT INTO products (name, stock) VALUES (%s, 0)", (name,))
        conn.commit()
        release_conn(conn)
        return redirect("/stock")

    if request.method == "POST" and request.form.get("form_type") == "delete_product":
        product_id = request.form["product_id"]
        cursor.execute("DELETE FROM products WHERE id=%s", (product_id,))
        conn.commit()
        release_conn(conn)
        return redirect("/stock")

    cursor.execute("SELECT id, name, COALESCE(stock, 0) FROM products")
    products = cursor.fetchall()
    release_conn(conn)
    return render_template("stock.html", products=products)


# ============================================================
# DELETE PRODUCT
# FIX: ? → %s
# ============================================================

@app.route("/delete_product/<int:id>")
@login_required
def delete_product(id):
    conn = connect_db()
    cursor = conn.cursor()
    cursor.execute("DELETE FROM products WHERE id=%s", (id,))
    conn.commit()
    release_conn(conn)
    return redirect("/stock")


# ============================================================
# DIESEL EXPENSES
# FIX: ? → %s, DATE('now') → CURRENT_DATE::TEXT
# ============================================================

@app.route("/diesel", methods=["GET", "POST"])
@login_required
def diesel():
    conn = connect_db()
    cursor = conn.cursor()

    if request.method == "POST":
        amount = int(request.form["amount"])
        note = request.form.get("note", "")
        cursor.execute("""
            INSERT INTO diesel_expenses (amount, note, date) VALUES (%s, %s, CURRENT_DATE::TEXT)
        """, (amount, note))
        conn.commit()
        release_conn(conn)
        return redirect("/diesel")

    cursor.execute("SELECT amount, note, date FROM diesel_expenses ORDER BY id DESC")
    data = cursor.fetchall()
    release_conn(conn)
    return render_template("diesel.html", data=data)


# ============================================================
# EXPENSE REPORT
# FIX: ? → %s
# ============================================================

@app.route("/expense_report", methods=["GET", "POST"])
@login_required
def expense_report():
    conn = connect_db()
    cursor = conn.cursor()

    from_date = request.form.get("from_date")
    to_date = request.form.get("to_date")

    labour_total = 0
    transport_total = 0
    diesel_total = 0

    if from_date and to_date:
        cursor.execute("""
            SELECT COALESCE(SUM(labour),0), COALESCE(SUM(transport),0)
            FROM sales WHERE date BETWEEN %s AND %s
        """, (from_date, to_date))
        data = cursor.fetchone()
        labour_total = data[0]
        transport_total = data[1]

        cursor.execute("""
            SELECT COALESCE(SUM(amount),0) FROM diesel_expenses WHERE date BETWEEN %s AND %s
        """, (from_date, to_date))
        diesel_total = cursor.fetchone()[0]

    release_conn(conn)
    return render_template("expense_report.html",
        labour=labour_total,
        transport=transport_total,
        diesel=diesel_total,
        from_date=from_date,
        to_date=to_date
    )


# ============================================================
# SALES REPORT PAGES
# FIX: DATE('now') → CURRENT_DATE::TEXT, strftime → TO_CHAR
# FIX: GROUP_CONCAT → STRING_AGG
# ============================================================

@app.route("/today_sales")
@login_required
def today_sales_page():
    conn = connect_db()
    cursor = conn.cursor()
    cursor.execute("""
        SELECT 
            s.id,
            c.name,
            STRING_AGG(si.item || '(' || si.quantity::TEXT || ')', ', ') AS items,
            STRING_AGG(si.quantity::TEXT, ', ') AS qtys,
            s.total,
            s.paid,
            s.pending,
            s.date
        FROM sales s
        JOIN customers c ON s.customer_id = c.id
        LEFT JOIN sale_items si ON s.id = si.sale_id
        WHERE s.date = CURRENT_DATE::TEXT
        GROUP BY s.id, c.name, s.total, s.paid, s.pending, s.date
        ORDER BY s.id DESC
    """)
    data = cursor.fetchall()
    release_conn(conn)
    return render_template("sales_report_page.html", title="Today's Sales", data=data)


@app.route("/monthly_sales")
@login_required
def monthly_sales_page():
    conn = connect_db()
    cursor = conn.cursor()
    cursor.execute("""
        SELECT 
            s.id,
            c.name,
            STRING_AGG(si.item || '(' || si.quantity::TEXT || ')', ', ') AS items,
            STRING_AGG(si.quantity::TEXT, ', ') AS qtys,
            s.total,
            s.paid,
            s.pending,
            s.date
        FROM sales s
        JOIN customers c ON s.customer_id = c.id
        LEFT JOIN sale_items si ON s.id = si.sale_id
        WHERE TO_CHAR(s.date::DATE, 'YYYY-MM') = TO_CHAR(CURRENT_DATE, 'YYYY-MM')
        GROUP BY s.id, c.name, s.total, s.paid, s.pending, s.date
        ORDER BY s.id DESC
    """)
    data = cursor.fetchall()
    release_conn(conn)
    return render_template("sales_report_page.html", title="Monthly Sales", data=data)


@app.route("/all_sales")
@login_required
def all_sales_page():
    conn = connect_db()
    cursor = conn.cursor()
    cursor.execute("""
        SELECT 
            s.id,
            c.name,
            STRING_AGG(si.item || '(' || si.quantity::TEXT || ')', ', ') AS items,
            STRING_AGG(si.quantity::TEXT, ', ') AS qtys,
            s.total,
            s.paid,
            s.pending,
            s.date
        FROM sales s
        JOIN customers c ON s.customer_id = c.id
        LEFT JOIN sale_items si ON s.id = si.sale_id
        GROUP BY s.id, c.name, s.total, s.paid, s.pending, s.date
        ORDER BY s.id DESC
    """)
    data = cursor.fetchall()
    release_conn(conn)
    return render_template("sales_report_page.html", title="All Sales", data=data)


# ============================================================
# COLLECTIONS REPORT PAGES
# FIX: DATE('now') → CURRENT_DATE::TEXT, strftime → TO_CHAR
# ============================================================

@app.route("/today_collections")
@login_required
def today_collections_page():
    conn = connect_db()
    cursor = conn.cursor()
    cursor.execute("""
        SELECT p.id, c.name, p.amount, p.mode, p.reference, p.notes, p.date
        FROM payments p JOIN customers c ON p.customer_id = c.id
        WHERE p.date = CURRENT_DATE::TEXT

        UNION ALL

        SELECT s.id, c.name, s.paid, 'Sale Payment', '', 'Paid during sale', s.date
        FROM sales s JOIN customers c ON s.customer_id = c.id
        WHERE s.date = CURRENT_DATE::TEXT AND s.paid > 0

        ORDER BY date DESC
    """)
    data = cursor.fetchall()
    release_conn(conn)
    return render_template("collections_report_page.html", title="Today's Collections", data=data)


@app.route("/monthly_collections")
@login_required
def monthly_collections_page():
    conn = connect_db()
    cursor = conn.cursor()
    cursor.execute("""
        SELECT p.id, c.name, p.amount, p.mode, p.reference, p.notes, p.date
        FROM payments p JOIN customers c ON p.customer_id = c.id
        WHERE TO_CHAR(p.date::DATE, 'YYYY-MM') = TO_CHAR(CURRENT_DATE, 'YYYY-MM')

        UNION ALL

        SELECT s.id, c.name, s.paid, 'Sale Payment', '', 'Paid during sale', s.date
        FROM sales s JOIN customers c ON s.customer_id = c.id
        WHERE TO_CHAR(s.date::DATE, 'YYYY-MM') = TO_CHAR(CURRENT_DATE, 'YYYY-MM') AND s.paid > 0

        ORDER BY date DESC
    """)
    data = cursor.fetchall()
    release_conn(conn)
    return render_template("collections_report_page.html", title="Monthly Collections", data=data)


@app.route("/all_collections")
@login_required
def all_collections_page():
    conn = connect_db()
    cursor = conn.cursor()
    cursor.execute("""
        SELECT p.id, c.name, p.amount, p.mode, p.reference, p.notes, p.date
        FROM payments p JOIN customers c ON p.customer_id = c.id

        UNION ALL

        SELECT s.id, c.name, s.paid, 'Sale Payment', '', 'Paid during sale', s.date
        FROM sales s JOIN customers c ON s.customer_id = c.id
        WHERE s.paid > 0

        ORDER BY date DESC
    """)
    data = cursor.fetchall()
    release_conn(conn)
    return render_template("collections_report_page.html", title="All Collections", data=data)


# ============================================================
# PENDING REPORT PAGES
# FIX: DATE('now') → CURRENT_DATE::TEXT, strftime → TO_CHAR
# FIX: HAVING with alias → wrap in subquery (PostgreSQL doesn't allow HAVING on SELECT aliases)
# ============================================================
@app.route("/today_pending")
@login_required
def today_pending_page():
    conn = connect_db()
    cursor = conn.cursor()
    cursor.execute("""
        SELECT * FROM (
            SELECT
                c.id,
                c.name,
                COALESCE(SUM(s.total), 0) AS total_amount,
                COALESCE(SUM(s.paid), 0) AS sale_paid,
                COALESCE((SELECT SUM(p.amount) FROM payments p WHERE p.customer_id = c.id), 0) AS payment_paid,
                (
                    COALESCE(c.opening_balance, 0)
                    + COALESCE((SELECT SUM(s2.total) FROM sales s2 WHERE s2.customer_id = c.id), 0)
                    - COALESCE((SELECT SUM(s2.paid) FROM sales s2 WHERE s2.customer_id = c.id), 0)
                    - COALESCE((SELECT SUM(p.amount) FROM payments p WHERE p.customer_id = c.id), 0)
                ) AS final_pending,
                MAX(s.date) AS date
            FROM customers c
            JOIN sales s ON s.customer_id = c.id
            WHERE s.date = CURRENT_DATE::TEXT
            GROUP BY c.id, c.name, c.opening_balance
        ) sub
        WHERE final_pending != 0
        ORDER BY final_pending DESC
    """)
    data = cursor.fetchall()
    release_conn(conn)
    return render_template("pending_report_page.html", title="Today's Pending", data=data)


@app.route("/monthly_pending")
@login_required
def monthly_pending_page():
    conn = connect_db()
    cursor = conn.cursor()
    cursor.execute("""
        SELECT * FROM (
            SELECT
                c.id,
                c.name,
                COALESCE(SUM(s.total), 0) AS total_amount,
                COALESCE(SUM(s.paid), 0) AS sale_paid,
                COALESCE((SELECT SUM(p.amount) FROM payments p WHERE p.customer_id = c.id), 0) AS payment_paid,
                (
                    COALESCE(c.opening_balance, 0)
                    + COALESCE((SELECT SUM(s2.total) FROM sales s2 WHERE s2.customer_id = c.id), 0)
                    - COALESCE((SELECT SUM(s2.paid) FROM sales s2 WHERE s2.customer_id = c.id), 0)
                    - COALESCE((SELECT SUM(p.amount) FROM payments p WHERE p.customer_id = c.id), 0)
                ) AS final_pending,
                MAX(s.date) AS date
            FROM customers c
            JOIN sales s ON s.customer_id = c.id
            WHERE TO_CHAR(s.date::DATE, 'YYYY-MM') = TO_CHAR(CURRENT_DATE, 'YYYY-MM')
            GROUP BY c.id, c.name, c.opening_balance
        ) sub
        WHERE final_pending != 0
        ORDER BY final_pending DESC
    """)
    data = cursor.fetchall()
    release_conn(conn)
    return render_template("pending_report_page.html", title="Monthly Pending", data=data)


@app.route("/all_pending")
@login_required
def all_pending_page():
    conn = connect_db()
    cursor = conn.cursor()
    cursor.execute("""
        SELECT * FROM (
            SELECT
                c.id,
                c.name,
                COALESCE((SELECT SUM(s.total) FROM sales s WHERE s.customer_id = c.id), 0)
                + COALESCE(c.opening_balance, 0) AS total_amount,

                COALESCE((SELECT SUM(s.paid) FROM sales s WHERE s.customer_id = c.id), 0) AS sale_paid,

                COALESCE((SELECT SUM(p.amount) FROM payments p WHERE p.customer_id = c.id), 0) AS payment_paid,

                (
                    COALESCE(c.opening_balance, 0)
                    + COALESCE((SELECT SUM(s.total) FROM sales s WHERE s.customer_id = c.id), 0)
                    - COALESCE((SELECT SUM(s.paid) FROM sales s WHERE s.customer_id = c.id), 0)
                    - COALESCE((SELECT SUM(p.amount) FROM payments p WHERE p.customer_id = c.id), 0)
                ) AS final_pending,

                COALESCE(
                    (SELECT MAX(s.date) FROM sales s WHERE s.customer_id = c.id),
                    CURRENT_DATE::TEXT
                ) AS date
            FROM customers c
        ) sub
        WHERE final_pending != 0
        ORDER BY final_pending DESC
    """)
    data = cursor.fetchall()
    release_conn(conn)
    return render_template("pending_report_page.html", title="All Pending", data=data)

# ============================================================
# RECALCULATE CUSTOMER PENDING
# FIX: ? → %s
# ============================================================

def recalculate_customer_pending(customer_id):
    customer_id = int(customer_id)
    conn = connect_db()
    cursor = conn.cursor()

    cursor.execute("UPDATE sales SET pending = total - paid WHERE customer_id=%s", (customer_id,))

    cursor.execute("SELECT opening_balance FROM customers WHERE id=%s", (customer_id,))
    row = cursor.fetchone()
    opening_balance = row[0] if row else 0

    cursor.execute("""
        SELECT amount FROM payments WHERE customer_id=%s ORDER BY date ASC, id ASC
    """, (customer_id,))
    payments = cursor.fetchall()

    cursor.execute("""
        SELECT id, pending FROM sales WHERE customer_id=%s AND pending > 0 ORDER BY date ASC, id ASC
    """, (customer_id,))
    sales = cursor.fetchall()
    sales_list = [[s[0], s[1]] for s in sales]

    for (amount,) in payments:
        pay_amount = amount

        if opening_balance > 0:
            if pay_amount >= opening_balance:
                pay_amount -= opening_balance
                opening_balance = 0
            else:
                opening_balance -= pay_amount
                pay_amount = 0

        if pay_amount <= 0:
            continue

        for sale in sales_list:
            if pay_amount <= 0:
                break
            sale_id = sale[0]
            pending = sale[1]
            if pending > 0:
                if pay_amount >= pending:
                    pay_amount -= pending
                    sale[1] = 0
                else:
                    sale[1] -= pay_amount
                    pay_amount = 0

    for sale in sales_list:
        cursor.execute("UPDATE sales SET pending = %s WHERE id = %s", (sale[1], sale[0]))

    conn.commit()
    release_conn(conn)


# ============================================================
# RECALCULATE SUPPLIER PENDING
# FIX: ? → %s
# ============================================================

def recalculate_supplier_pending(supplier_id):
    supplier_id = int(supplier_id)
    conn = connect_db()
    cursor = conn.cursor()

    cursor.execute("UPDATE purchases SET pending = total - paid WHERE supplier_id=%s", (supplier_id,))

    cursor.execute("""
        SELECT amount FROM supplier_payments WHERE supplier_id=%s ORDER BY date ASC, id ASC
    """, (supplier_id,))
    payments = cursor.fetchall()

    cursor.execute("""
        SELECT id, pending FROM purchases WHERE supplier_id=%s AND pending > 0 ORDER BY date ASC, id ASC
    """, (supplier_id,))
    purchases = cursor.fetchall()
    purchases_list = [[p[0], p[1]] for p in purchases]

    for (amount,) in payments:
        pay_amount = amount
        for purchase in purchases_list:
            if pay_amount <= 0:
                break
            purchase_id = purchase[0]
            pending = purchase[1]
            if pending > 0:
                if pay_amount >= pending:
                    pay_amount -= pending
                    purchase[1] = 0
                else:
                    purchase[1] -= pay_amount
                    pay_amount = 0

    for purchase in purchases_list:
        cursor.execute("UPDATE purchases SET pending = %s WHERE id = %s", (purchase[1], purchase[0]))

    conn.commit()
    release_conn(conn)


# ============================================================
# FIX ALL PENDING
# ============================================================

@app.route("/fix_all_pending")
@login_required
def fix_all_pending():
    conn = connect_db()
    cursor = conn.cursor()
    cursor.execute("SELECT id FROM customers")
    customers = cursor.fetchall()
    release_conn(conn)

    for (customer_id,) in customers:
        recalculate_customer_pending(customer_id)

    return "✅ All pending amounts recalculated successfully!"


# ============================================================
# DOWNLOAD ALL PENDING (EXCEL)
# FIX: GROUP_CONCAT → STRING_AGG, ? → %s
# FIX: pd.read_sql_query → cursor + DataFrame
# ============================================================

@app.route("/download_all_pending")
@login_required
def download_all_pending():
    conn = connect_db()
    cursor = conn.cursor()

    cursor.execute("""
        SELECT 
            c.id AS ID,
            c.name AS Customer,
            'Opening Balance' AS Items,
            c.opening_balance AS Total,
            0 AS Sale_Paid,
            0 AS Payment,
            c.opening_balance AS Final_Pending,
            '-' AS Date
        FROM customers c
        WHERE c.opening_balance != 0

        UNION ALL

        SELECT 
            s.id,
            c.name,
            COALESCE(STRING_AGG(si.item || '(' || si.quantity::TEXT || ')', ', '), ''),
            s.total,
            s.paid,
            COALESCE((SELECT SUM(p.amount) FROM payments p WHERE p.customer_id = s.customer_id), 0),
            s.pending - COALESCE((SELECT SUM(p.amount) FROM payments p WHERE p.customer_id = s.customer_id), 0),
            s.date
        FROM sales s
        JOIN customers c ON s.customer_id = c.id
        LEFT JOIN sale_items si ON s.id = si.sale_id
        WHERE s.pending != 0
        GROUP BY s.id, c.name, s.total, s.paid, s.pending, s.date, s.customer_id

        ORDER BY Final_Pending DESC
    """)
    rows = cursor.fetchall()
    release_conn(conn)

    columns = ["ID", "Customer", "Items", "Total", "Sale Paid", "Payment", "Final Pending", "Date"]
    df = pd.DataFrame(rows, columns=columns)

    output = io.BytesIO()
    df.to_excel(output, index=False)
    output.seek(0)

    return send_file(output, download_name="all_pending_report.xlsx", as_attachment=True)


# ============================================================
# DOWNLOAD ALL PENDING (PDF)
# FIX: GROUP_CONCAT → STRING_AGG
# ============================================================

@app.route("/download_all_pending_pdf")
@login_required
def download_all_pending_pdf():
    conn = connect_db()
    cursor = conn.cursor()

    cursor.execute("""
        SELECT 
            c.id,
            c.name,
            'Opening Balance',
            c.opening_balance,
            0,
            0,
            c.opening_balance,
            '-'
        FROM customers c
        WHERE c.opening_balance != 0

        UNION ALL

        SELECT 
            s.id,
            c.name,
            COALESCE(STRING_AGG(si.item || '(' || si.quantity::TEXT || ')', ', '), ''),
            s.total,
            s.paid,
            COALESCE((SELECT SUM(p.amount) FROM payments p WHERE p.customer_id = s.customer_id), 0),
            s.pending - COALESCE((SELECT SUM(p.amount) FROM payments p WHERE p.customer_id = s.customer_id), 0),
            s.date
        FROM sales s
        JOIN customers c ON s.customer_id = c.id
        LEFT JOIN sale_items si ON s.id = si.sale_id
        WHERE s.pending != 0
        GROUP BY s.id, c.name, s.total, s.paid, s.pending, s.date, s.customer_id

        ORDER BY 7 DESC
    """)
    rows = cursor.fetchall()
    release_conn(conn)

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4))
    elements = []
    styles = getSampleStyleSheet()

    elements.append(Paragraph("<b>All Pending Report</b>", styles["Title"]))
    elements.append(Spacer(1, 12))

    data = [["ID", "Customer", "Items", "Total", "Sale Paid", "Payment", "Final Pending", "Date"]]
    for r in rows:
        data.append([r[0], r[1], r[2], f"₹{r[3]}", f"₹{r[4]}", f"₹{r[5]}", f"₹{r[6]}", r[7]])

    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 7),
        ("ALIGN", (3, 1), (-2, -1), "RIGHT"),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
    ]))
    elements.append(table)
    doc.build(elements)
    buffer.seek(0)

    return send_file(buffer, as_attachment=True, download_name="all_pending_report.pdf", mimetype="application/pdf")


# ============================================================
# RUN
# ============================================================

if __name__ == "__main__":
    app.run(debug=True)